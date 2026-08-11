import openpyxl
wb = openpyxl.load_workbook(r'E:\webdev\ANC_Conf_Nominations\docs\NOM2026 PR and Councillor Nominations.xlsx', data_only=True)
ws = wb['TOTAL IN ZONES']

print('Per-zone breakdown from TOTAL IN ZONES:')
print(f'{"Candidate":<25} {"AMON":>5} {"ANDREW":>6} {"PAUL":>5} {"BAVUMILE":>8} {"GRACE":>6} {"ZONE10":>7} {"SELOPE":>6} {"DAVID":>6} {"DIZA":>5} {"OSKA":>5} {"TOTAL":>6}')
print('-' * 95)

for row_idx in range(6, 40):
    name = ws.cell(row=row_idx, column=1).value
    if name and str(name).strip() and str(name).strip() != 'OTHERS':
        vals = []
        for col in range(2, 13):
            v = ws.cell(row=row_idx, column=col).value
            vals.append(v if v else 0)
        name_str = str(name).strip()
        print(f'{name_str:<25} {vals[0]:>5} {vals[1]:>6} {vals[2]:>5} {vals[3]:>8} {vals[4]:>6} {vals[5]:>7} {vals[6]:>6} {vals[7]:>6} {vals[8]:>5} {vals[9]:>5} {vals[10]:>6}')
