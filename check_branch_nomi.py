import openpyxl
from collections import defaultdict
import re

wb = openpyxl.load_workbook(r'E:\webdev\ANC_Conf_Nominations\docs\NOM2026 PR and Councillor Nominations.xlsx', data_only=True)
ws = wb['BRANCH NOMI']

print(f"BRANCH NOMI dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
print()

# Check header row (row 3)
print("Row 3 (headers):")
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=3, column=col).value
    if val:
        print(f"  Col {col}: {val}")

print()

# Count data rows
data_rows = 0
for row_idx in range(4, ws.max_row + 1):
    name = ws.cell(row=row_idx, column=1).value
    if name:
        data_rows += 1

print(f"Data rows (from col A): {data_rows}")

# Extract per-zone vote counts from BRANCH NOMI
# Based on seed_data.py logic - it counts occurrences of each name per zone/ward
zone_ward_candidates = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

WARD_PATTERN = re.compile(r"WARD\s*(\d+)", re.IGNORECASE)
header_row = 3

# Get zone labels from header row
zone_labels = {}
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=header_row, column=col_idx).value
    if val:
        zone_labels[col_idx] = str(val).strip()

print(f"\nZone columns: {zone_labels}")

# Parse like seed_data.py does
for col_idx, zone_label in zone_labels.items():
    if not zone_label.lower().startswith("column"):
        current_ward = None
        for row_idx in range(header_row + 1, ws.max_row + 1):
            raw_value = str(ws.cell(row=row_idx, column=col_idx).value or "").strip()
            if not raw_value:
                continue

            ward_match = WARD_PATTERN.search(raw_value)
            if ward_match:
                current_ward = int(ward_match.group(1))
                continue

            if current_ward is None:
                continue

            if raw_value.upper().startswith("WARD"):
                continue

            # This is a candidate name - count as 1 vote
            zone_ward_candidates[zone_label][current_ward][raw_value] += 1

# Now tally totals per zone and per candidate
branch_zone_totals = defaultdict(int)
branch_candidate_totals = defaultdict(int)
total_votes = 0

for zone, wards in zone_ward_candidates.items():
    for ward, candidates in wards.items():
        for candidate, votes in candidates.items():
            branch_zone_totals[zone] += votes
            branch_candidate_totals[candidate] += votes
            total_votes += votes

print(f"\n=== BRANCH NOMI totals ===")
print(f"Total votes (from branch data): {total_votes}")
print()
print("Zone totals:")
for zone, total in sorted(branch_zone_totals.items()):
    print(f"  {zone}: {total}")

print()
print("Candidate totals:")
for name, total in sorted(branch_candidate_totals.items(), key=lambda x: -x[1]):
    print(f"  {name}: {total}")
