"""Check if TOTAL row is a formula and verify Excel calculations."""
import openpyxl

# Load with formulas (not data_only)
wb = openpyxl.load_workbook(r'E:\webdev\ANC_Conf_Nominations\docs\NOM2026 PR and Councillor Nominations.xlsx')
ws = wb['TOTAL IN ZONES']

print("TOTAL row (row 51) - checking formulas:")
for col_idx in range(1, 13):
    cell = ws.cell(row=51, column=col_idx)
    print(f"  Col {col_idx} ({cell.coordinate}): value={cell.value}, type={type(cell.value).__name__}")

print("\nSample data cells - checking if formulas:")
for row_idx in [6, 7, 19]:
    for col_idx in [2, 3, 12]:
        cell = ws.cell(row=row_idx, column=col_idx)
        print(f"  {cell.coordinate}: value={cell.value}, type={type(cell.value).__name__}")

# Now check with data_only
wb2 = openpyxl.load_workbook(r'E:\webdev\ANC_Conf_Nominations\docs\NOM2026 PR and Councillor Nominations.xlsx', data_only=True)
ws2 = wb2['TOTAL IN ZONES']

# Sum all candidate votes per zone
zone_sums = {}
for col_idx in range(2, 12):
    zone_name = ws2.cell(row=4, column=col_idx).value
    if zone_name:
        total = 0
        for row_idx in range(6, 51):
            val = ws2.cell(row=row_idx, column=col_idx).value
            if val and str(val).strip():
                name = ws2.cell(row=row_idx, column=1).value
                if name and str(name).strip() not in ('OTHERS', 'TOTAL', ''):
                    total += int(val)
        zone_sums[str(zone_name).strip()] = total

print("\n\nZone sums (from individual candidates):")
grand = 0
for zone, total in sorted(zone_sums.items()):
    row51_val = ws2.cell(row=51, column=[k for k, v in {2: 'AMON NGULELE', 3: 'ANDREW MAPHETO', 4: 'PAUL SEHLOHO', 5: 'BAVUMILE VILAKAZI', 6: 'GRACE FLATELA', 7: 'ZONE 10', 8: 'SELOPE THEMA', 9: 'DAVID BOPAPE', 10: 'DIZA PHUTHINI', 11: 'OSKA MABIKA'}.items() if v == zone][0]).value if zone in {2: 'AMON NGULELE', 3: 'ANDREW MAPHETO', 4: 'PAUL SEHLOHO', 5: 'BAVUMILE VILAKAZI', 6: 'GRACE FLATELA', 7: 'ZONE 10', 8: 'SELOPE THEMA', 9: 'DAVID BOPAPE', 10: 'DIZA PHUTHINI', 11: 'OSKA MABIKA'}.values() else 0
    # Get the column index for this zone
    col_map = {'AMON NGULELE': 2, 'ANDREW MAPHETO': 3, 'PAUL SEHLOHO': 4, 'BAVUMILE VILAKAZI': 5, 'GRACE FLATELA': 6, 'ZONE 10': 7, 'SELOPE THEMA': 8, 'DAVID BOPAPE': 9, 'DIZA PHUTHINI': 10, 'OSKA MABIKA': 11}
    row51 = ws2.cell(row=51, column=col_map.get(zone, 0)).value if zone in col_map else 0
    diff = total - (int(row51) if row51 else 0)
    print(f"  {zone}: sum={total}, TOTAL_row={row51}, diff={diff}")
    grand += total

print(f"\nGrand total from individual candidates: {grand}")
print(f"TOTAL row L51: {ws2.cell(row=51, column=12).value}")
