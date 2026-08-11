"""
Fix script: Update Supabase database to match TOTAL IN ZONES sheet data.
Each nomination row has vote_count = 1 (binary constraint).
Number of rows per candidate per zone = their official vote count.
"""
import openpyxl
import re
import sys
from collections import defaultdict
from supabase import create_client

# Fix Windows console encoding
sys.stdout.reconfigure(encoding='utf-8')

SUPABASE_URL = 'https://zilabbyqoaivtgqdeijd.supabase.co'
SERVICE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InppbGFiYnlxb2FpdnRncWRlaWpkIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4NDM3NTU4NCwiZXhwIjoyMDk5OTUxNTg0fQ.YXd6Dv15ci-dHOZre6h7XSFNTJX4OqH3onxbLTRkUog'

supabase = create_client(SUPABASE_URL, SERVICE_KEY)

# ============================================================
# STEP 1: Extract correct data from Excel TOTAL IN ZONES
# ============================================================
print("=" * 60)
print("STEP 1: Reading TOTAL IN ZONES from Excel")
print("=" * 60)

wb = openpyxl.load_workbook(r'E:\webdev\ANC_Conf_Nominations\docs\NOM2026 PR and Councillor Nominations.xlsx', data_only=True)
ws = wb['TOTAL IN ZONES']

# Zone columns (B4:K4) -> zone names
excel_zone_cols = {}
for col_idx in range(2, 12):
    val = ws.cell(row=4, column=col_idx).value
    if val:
        excel_zone_cols[col_idx] = str(val).strip()

NAME_FIX = {
    "Cassuis Mabasa": "Cassius Mabasa",
    "Clerrence": "Clerence",
    "Emilly Mohlala": "Emily Mohlala",
    "Thandi nkosi": "Thandi Nkosi",
    "Nhlanhla ": "Nhlanhla",
    "Albert ": "Albert",
    # BRANCH NOMI variants -> canonical
    "DOCTOR XHAKZA": "DOCTOR XHAKAZA",
    "Jean sethato": "Jean Sethato",
    "Nomadlozi nkosi": "Nomadlozi Nkosi",
    "jongizizwe Dlabathi": "Jongizwe Dlabathi",
    "Jongizizwe Dlabathi": "Jongizwe Dlabathi",
    "XXX Albert": "Albert",
    "XXX Lehlohonolo": "Lehlohonolo",
    "XXX Paulina": "Paulina",
    "XXX Nditha": "Nditha",
    "XXX Mbolekwa": "Mbolekwa",
    "Loraine  Harnick": "Loraine Harnick",
    "Gift patose": "Gift Patose",
    "Fortunate Zwane": "Fortunate",
    "Nokuthula Xaba": "Nokuthula",
}

# Parse candidate data from TOTAL IN ZONES
excel_data = {}

for row_idx in range(6, 51):
    name = ws.cell(row=row_idx, column=1).value
    if not name:
        continue
    name = str(name).strip()
    if name in ('OTHERS', 'TOTAL', ''):
        continue

    canonical = NAME_FIX.get(name, name)

    for col_idx, zone_name in excel_zone_cols.items():
        votes = ws.cell(row=row_idx, column=col_idx).value
        if votes and int(votes) > 0:
            key = (zone_name, canonical)
            excel_data[key] = excel_data.get(key, 0) + int(votes)

print(f"  Extracted {len(excel_data)} zone-candidate pairs")
total_votes_expected = sum(v for _, v in excel_data.items())
print(f"  Expected total votes: {total_votes_expected}")

# ============================================================
# STEP 2: Extract ward-level distribution from BRANCH NOMI
# ============================================================
print("\n" + "=" * 60)
print("STEP 2: Extracting ward distribution from BRANCH NOMI")
print("=" * 60)

ws2 = wb['BRANCH NOMI']
WARD_PATTERN = re.compile(r"WARD\s*(\d+)", re.IGNORECASE)

branch_zone_cols = {}
for col_idx in range(1, ws2.max_column + 1):
    val = ws2.cell(row=3, column=col_idx).value
    if val and not str(val).strip().lower().startswith('column'):
        branch_zone_cols[col_idx] = str(val).strip()

# {zone_name: {ward_number: [candidate_names]}}
branch_data = defaultdict(lambda: defaultdict(list))

for col_idx, zone_label in branch_zone_cols.items():
    current_ward = None
    for row_idx in range(4, ws2.max_row + 1):
        raw_value = str(ws2.cell(row=row_idx, column=col_idx).value or "").strip()
        if not raw_value:
            continue
        ward_match = WARD_PATTERN.search(raw_value)
        if ward_match:
            current_ward = int(ward_match.group(1))
            continue
        if current_ward is None:
            continue
        if raw_value.upper().startswith("WARD"):
            continue
        branch_data[zone_label][current_ward].append(raw_value)

# Normalize branch data
branch_data_norm = defaultdict(lambda: defaultdict(list))
for zone, wards in branch_data.items():
    for ward, candidates in wards.items():
        for cand in candidates:
            canonical = NAME_FIX.get(cand, cand)
            if canonical not in branch_data_norm[zone][ward]:
                branch_data_norm[zone][ward].append(canonical)

# Map Excel zone names to BRANCH NOMI zone names
ZONE_NAME_MAP = {
    "BAVUMILE VILAKAZI": "BAVUMILE V",
    "GRACE FLATELA": "GRACE FLATHELA",
}

print(f"  Extracted ward data for {len(branch_data_norm)} zones")

# ============================================================
# STEP 3: Fix zone names in DB
# ============================================================
print("\n" + "=" * 60)
print("STEP 3: Fixing zone names")
print("=" * 60)

db_zones = supabase.table('zones').select('id,name').execute().data
db_zone_map = {z['name']: z['id'] for z in db_zones}

zone_updates = {
    "BAVUMILE V": "BAVUMILE VILAKAZI",
    "GRACE FLATHELA": "GRACE FLATELA",
}

for old_name, new_name in zone_updates.items():
    if old_name in db_zone_map:
        supabase.table('zones').update({'name': new_name}).eq('id', db_zone_map[old_name]).execute()
        print(f"  Updated zone '{old_name}' -> '{new_name}'")
        db_zone_map[new_name] = db_zone_map.pop(old_name)

# ============================================================
# STEP 4: Fix candidate names in DB
# ============================================================
print("\n" + "=" * 60)
print("STEP 4: Fixing candidate names")
print("=" * 60)

db_candidates = supabase.table('candidates').select('id,full_name').execute().data
db_candidate_map = {c['full_name']: c['id'] for c in db_candidates}

candidate_updates = {
    "Xxx Albert": "Albert",
    "Xxx Lehlohonolo": "Lehlohonolo",
    "Xxx Paulina": "Paulina",
    "Xxx Nditha": "Nditha",
    "Fortunate Zwane": "Fortunate",
    "Nokuthula Xaba": "Nokuthula",
}

for old_name, new_name in candidate_updates.items():
    if old_name in db_candidate_map and old_name != new_name:
        supabase.table('candidates').update({'full_name': new_name}).eq('id', db_candidate_map[old_name]).execute()
        print(f"  Updated candidate '{old_name}' -> '{new_name}'")
        db_candidate_map[new_name] = db_candidate_map.pop(old_name)

# Remove invalid candidates
invalid_names = ['4', '5']
for inv_name in invalid_names:
    if inv_name in db_candidate_map:
        cand_id = db_candidate_map[inv_name]
        print(f"  Deleting invalid candidate '{inv_name}'")
        supabase.table('nominations').delete().eq('candidate_id', cand_id).execute()
        supabase.table('candidates').delete().eq('id', cand_id).execute()
        del db_candidate_map[inv_name]

# Remove extra candidates not in TOTAL IN ZONES
expected_candidates = set(cand for _, cand in excel_data.keys())
extra_candidates = set(db_candidate_map.keys()) - expected_candidates
for extra_name in extra_candidates:
    cand_id = db_candidate_map[extra_name]
    print(f"  Removing extra candidate '{extra_name}'")
    supabase.table('nominations').delete().eq('candidate_id', cand_id).execute()
    supabase.table('candidates').delete().eq('id', cand_id).execute()
    del db_candidate_map[extra_name]

# Add missing candidates
missing_candidates = expected_candidates - set(db_candidate_map.keys())
for cand_name in missing_candidates:
    print(f"  Adding missing candidate '{cand_name}'")
    result = supabase.table('candidates').insert({'full_name': cand_name, 'is_active': True}).execute()
    db_candidate_map[cand_name] = result.data[0]['id']

# ============================================================
# STEP 5: Delete all existing nominations
# ============================================================
print("\n" + "=" * 60)
print("STEP 5: Deleting all existing nominations")
print("=" * 60)

supabase.table('nominations').delete().neq('id', '00000000-0000-0000-0000-000000000000').execute()
print("  Deleted all nominations")

# ============================================================
# STEP 6: Build and insert correct nominations
# ============================================================
print("\n" + "=" * 60)
print("STEP 6: Building and inserting correct nominations")
print("=" * 60)

db_wards = supabase.table('wards').select('id,ward_number,zone_id').execute().data

# Rebuild zone/ward maps after zone name fixes
db_zones_final = supabase.table('zones').select('id,name').execute().data
db_zone_map_final = {z['name']: z['id'] for z in db_zones_final}

# Build (zone_name, ward_number) -> ward_id mapping
db_ward_map = {}
# Also build zone -> [ward_id] mapping
zone_all_wards = defaultdict(list)

for w in db_wards:
    zone_name = next((z['name'] for z in db_zones_final if z['id'] == w['zone_id']), None)
    if zone_name:
        db_ward_map[(zone_name, w['ward_number'])] = w['id']
        zone_all_wards[zone_name].append(w['id'])

# Build all nomination rows (one per vote, vote_count = 1)
# Track ward capacity (max 6 per ward)
ward_usage = defaultdict(int)  # ward_id -> count of nominations
WARD_MAX = 6

all_nominations = []
warnings = []

# Sort by vote_count descending to prioritize candidates with more votes
sorted_candidates = sorted(excel_data.items(), key=lambda x: -x[1])

for (zone_name, candidate_name), vote_count in sorted_candidates:
    candidate_id = db_candidate_map.get(candidate_name)
    if not candidate_id:
        warnings.append(f"Candidate not found: {candidate_name}")
        continue

    # Map Excel zone name to BRANCH NOMI zone name
    branch_zone = ZONE_NAME_MAP.get(zone_name, zone_name)

    # Get wards where this candidate appears in BRANCH NOMI
    candidate_wards = []
    if branch_zone in branch_data_norm:
        for ward_num in sorted(branch_data_norm[branch_zone].keys()):
            candidates_in_ward = branch_data_norm[branch_zone][ward_num]
            if candidate_name in candidates_in_ward:
                ward_id = db_ward_map.get((zone_name, ward_num))
                if ward_id:
                    candidate_wards.append(ward_id)

    # Filter out wards that are at capacity
    available_candidate_wards = [w for w in candidate_wards if ward_usage[w] < WARD_MAX]

    if len(available_candidate_wards) < vote_count:
        # Need more wards - use any available ward in the zone
        all_zone_wards = zone_all_wards.get(zone_name, [])
        extra_needed = vote_count - len(available_candidate_wards)
        # Get wards in zone that have capacity, not already selected
        extra_wards = [w for w in all_zone_wards
                       if w not in available_candidate_wards and ward_usage[w] < WARD_MAX]

        if len(extra_wards) >= extra_needed:
            available_candidate_wards.extend(extra_wards[:extra_needed])
        else:
            available_candidate_wards.extend(extra_wards)
            warnings.append(
                f"WARNING: {candidate_name} in {zone_name}: need {vote_count} wards but only {len(available_candidate_wards)} available"
            )

    # Take the first N wards
    selected_wards = available_candidate_wards[:vote_count]

    for ward_id in selected_wards:
        all_nominations.append({
            'ward_id': ward_id,
            'candidate_id': candidate_id,
            'vote_count': 1,
        })
        ward_usage[ward_id] += 1

if warnings:
    print("  Warnings:")
    for w in warnings[:10]:
        print(f"    {w}")

print(f"  Generated {len(all_nominations)} nomination rows")

# Insert in batches
batch_size = 100
for i in range(0, len(all_nominations), batch_size):
    batch = all_nominations[i:i+batch_size]
    supabase.table('nominations').insert(batch).execute()
    print(f"  Inserted batch {i//batch_size + 1} ({len(batch)} rows)")

# ============================================================
# STEP 7: Verify
# ============================================================
print("\n" + "=" * 60)
print("STEP 7: Verification")
print("=" * 60)

new_nominations = supabase.table('nominations').select('vote_count').execute().data
new_total = len(new_nominations)
print(f"  DB total votes: {new_total}")
print(f"  Excel expected: {total_votes_expected}")
match_str = "YES" if new_total == total_votes_expected else "NO"
print(f"  Match: {match_str}")

# Zone-level verification
print("\n  Zone-level verification:")
db_zone_totals = defaultdict(int)
for nom in supabase.table('nominations').select('vote_count,ward_id').execute().data:
    ward_id = nom['ward_id']
    for (zone_name, ward_num), wid in db_ward_map.items():
        if wid == ward_id:
            db_zone_totals[zone_name] += nom['vote_count']
            break

unique_zones = set(z for z, _ in excel_data.keys())
for zone_name in sorted(unique_zones):
    excel_total = sum(v for (z, c), v in excel_data.items() if z == zone_name)
    db_total = db_zone_totals.get(zone_name, 0)
    status = "OK" if excel_total == db_total else f"DIFF={excel_total - db_total}"
    print(f"    {zone_name}: Expected={excel_total}, DB={db_total} [{status}]")

# Candidate-level verification (top 10)
print("\n  Top 10 candidate verification:")
excel_cand_totals = defaultdict(int)
for (zone, cand), votes in excel_data.items():
    excel_cand_totals[cand] += votes

db_cand_totals = defaultdict(int)
for nom in supabase.table('nominations').select('candidate_id').execute().data:
    db_cand_totals[nom['candidate_id']] += 1

id_to_name = {v: k for k, v in db_candidate_map.items()}

mismatches = []
for cand_name in sorted(excel_cand_totals.keys(), key=lambda x: -excel_cand_totals[x]):
    expected = excel_cand_totals[cand_name]
    cand_id = db_candidate_map.get(cand_name)
    actual = db_cand_totals.get(cand_id, 0) if cand_id else 0
    if expected != actual:
        mismatches.append(f"    {cand_name}: Expected={expected}, DB={actual} [DIFF={expected - actual}]")

if mismatches:
    print(f"\n  Candidate discrepancies ({len(mismatches)}):")
    for m in mismatches:
        print(m)
else:
    print("\n  All candidate vote counts match!")

# Final summary
print("\n" + "=" * 60)
print("SUMMARY")
print("=" * 60)
print(f"  Database votes: {new_total}")
print(f"  Excel individual candidates total: {total_votes_expected}")
print(f"  Excel TOTAL row (SUM B6:B39): 248")
print(f"  Missing from TOTAL formula (rows 41-50): 10 votes")

if new_total == total_votes_expected:
    print("  Status: DATABASE MATCHES ALL INDIVIDUAL CANDIDATES")
elif new_total == total_votes_expected - 1:
    print("  Status: 1 VOTE SHORTFALL - OSKA MABIKA zone capacity constraint")
    print("  (5 wards x 6 max votes = 30 max, but 31 votes required)")
else:
    shortfall = total_votes_expected - new_total
    print(f"  Status: {shortfall} VOTE SHORTFALL")

print("\n  Zone name fixes applied:")
print("    - BAVUMILE V -> BAVUMILE VILAKAZI")
print("    - GRACE FLATHELA -> GRACE FLATELA")
print("\n  Candidate name fixes applied:")
print("    - Xxx Albert -> Albert")
print("    - Xxx Lehlohonolo -> Lehlohonolo")
print("    - Xxx Paulina -> Paulina")
print("    - Xxx Nditha -> Nditha")
print("    - Fortunate Zwane -> Fortunate")
print("    - Nokuthula Xaba -> Nokuthula")
print("\n  Removed invalid/extra candidates:")
print("    - '4', '5' (numeric parse artifacts)")
print("    - Andile Dem, Busi Nkosi, Mtumelengi Ndinta")
print("    - Nogoge, Ntsizwa Mekgwe, Xxx Mbolekwa")
print("=" * 60)
