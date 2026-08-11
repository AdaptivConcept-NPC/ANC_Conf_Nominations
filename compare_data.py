import openpyxl
import urllib.request
import json
from collections import defaultdict

SUPABASE_URL = 'https://zilabbyqoaivtgqdeijd.supabase.co'
SERVICE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InppbGFiYnlxb2FpdnRncWRlaWpkIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4NDM3NTU4NCwiZXhwIjoyMDk5OTUxNTg0fQ.YXd6Dv15ci-dHOZre6h7XSFNTJX4OqH3onxbLTRkUog'

headers = {
    'apikey': SERVICE_KEY,
    'Authorization': f'Bearer {SERVICE_KEY}',
    'Content-Type': 'application/json'
}

def supabase_get(table, params=''):
    req = urllib.request.Request(
        f'{SUPABASE_URL}/rest/v1/{table}?{params}',
        headers=headers
    )
    resp = urllib.request.urlopen(req)
    return json.loads(resp.read().decode())

# ---- 1. Parse Excel TOTAL IN ZONES ----
wb = openpyxl.load_workbook(r'E:\webdev\ANC_Conf_Nominations\docs\NOM2026 PR and Councillor Nominations.xlsx', data_only=True)
ws = wb['TOTAL IN ZONES']

# Headers in row 4: B=AMON NGULELE, C=ANDREW MAPHETO, D=PAUL SEHLOHO, E=BAVUMILE VILAKAZI,
# F=GRACE FLATELA, G=ZONE 10, H=SELOPE THEMA, I=DAVID BOPAPE, J=DIZA PHUTHINI, K=OSKA MABIKA
zones_from_excel = {}
for col_idx in range(2, 12):  # B to K
    val = ws.cell(row=4, column=col_idx).value
    if val:
        zones_from_excel[col_idx] = str(val).strip()

print("Zone columns from Excel:")
for col, name in zones_from_excel.items():
    print(f"  Col {col}: {name}")

# Parse candidate totals from Excel
excel_totals = {}
for row_idx in range(6, 51):  # rows 6-50 have candidate data
    name = ws.cell(row=row_idx, column=1).value
    total = ws.cell(row=row_idx, column=12).value  # Column L = TOTAL
    if name and total is not None:
        name_clean = str(name).strip()
        if name_clean and name_clean != 'OTHERS':
            excel_totals[name_clean] = int(total)

# Also parse TOTAL row (row 51)
totals_row = {}
for col_idx in range(2, 12):
    val = ws.cell(row=51, column=col_idx).value
    if val:
        zone_name = zones_from_excel.get(col_idx)
        if zone_name:
            totals_row[zone_name] = int(val)
excel_grand_total = ws.cell(row=51, column=12).value

print(f"\n=== EXCEL DATA ===")
print(f"Candidates: {len(excel_totals)}")
print(f"Grand total: {excel_grand_total}")
print(f"Zone totals: {totals_row}")
print()
for name, total in sorted(excel_totals.items(), key=lambda x: -x[1]):
    print(f"  {name}: {total}")

# ---- 2. Get DB data ----
candidates = supabase_get('candidates', 'select=id,full_name,is_active')
wards = supabase_get('wards', 'select=id,ward_number,zone_id')
zones = supabase_get('zones', 'select=id,name')
nominations = supabase_get('nominations', 'select=id,vote_count,ward_id,candidate_id')

candidate_map = {c['id']: c['full_name'] for c in candidates}
zone_map = {z['id']: z['name'] for z in zones}
ward_zone = {}
for w in wards:
    zone_name = zone_map.get(w['zone_id'], 'UNKNOWN')
    ward_zone[w['id']] = (w['ward_number'], zone_name)

# Aggregate by candidate
db_candidate_totals = defaultdict(int)
for nom in nominations:
    cand_name = candidate_map.get(nom['candidate_id'], 'UNKNOWN')
    db_candidate_totals[cand_name] += nom['vote_count']

# Aggregate by zone
db_zone_totals = defaultdict(int)
for nom in nominations:
    ward_id = nom['ward_id']
    _, zone_name = ward_zone.get(ward_id, (None, 'UNKNOWN'))
    db_zone_totals[zone_name] += nom['vote_count']

db_grand_total = sum(db_candidate_totals.values())

print(f"\n=== DATABASE DATA ===")
print(f"Candidates: {len(db_candidate_totals)}")
print(f"Grand total: {db_grand_total}")
print(f"Zone totals: {dict(db_zone_totals)}")
print()
for name, total in sorted(db_candidate_totals.items(), key=lambda x: -x[1]):
    print(f"  {name}: {total}")

# ---- 3. Compare ----
print(f"\n=== COMPARISON ===")
print(f"Excel grand total: {excel_grand_total}")
print(f"DB grand total: {db_grand_total}")
print(f"Difference: {int(excel_grand_total) - db_grand_total}")

# Candidates in Excel not in DB or different
print(f"\n--- Candidate differences ---")
all_names = set(list(excel_totals.keys()) + list(db_candidate_totals.keys()))

# Try to match names (normalize)
def normalize_name(n):
    return ' '.join(n.split()).lower().strip()

excel_norm = {normalize_name(k): (k, v) for k, v in excel_totals.items()}
db_norm = {normalize_name(k): (k, v) for k, v in db_candidate_totals.items()}

print("\nCandidates in Excel but missing or different in DB:")
for norm_name, (orig_name, excel_votes) in sorted(excel_norm.items()):
    if norm_name in db_norm:
        db_orig, db_votes = db_norm[norm_name]
        if excel_votes != db_votes:
            print(f"  MISMATCH: '{orig_name}' (Excel) vs '{db_orig}' (DB) - Excel: {excel_votes}, DB: {db_votes}, Diff: {excel_votes - db_votes}")
    else:
        # Try partial match
        found = False
        for db_norm_name, (db_orig, db_votes) in db_norm.items():
            if norm_name in db_norm_name or db_norm_name in norm_name:
                print(f"  POSSIBLE MATCH: Excel '{orig_name}' ({excel_votes}) ~ DB '{db_orig}' ({db_votes})")
                found = True
                break
        if not found:
            print(f"  MISSING IN DB: '{orig_name}' = {excel_votes} votes")

print("\nCandidates in DB but not in Excel TOTAL IN ZONES:")
for norm_name, (orig_name, db_votes) in sorted(db_norm.items()):
    if norm_name not in excel_norm:
        # Try partial match
        found = False
        for excel_norm_name, (excel_orig, excel_votes) in excel_norm.items():
            if norm_name in excel_norm_name or excel_norm_name in norm_name:
                found = True
                break
        if not found:
            print(f"  EXTRA IN DB: '{orig_name}' = {db_votes} votes")

# Zone comparison
print(f"\n--- Zone comparison ---")
for zone_name in sorted(set(list(totals_row.keys()) + list(db_zone_totals.keys()))):
    excel_val = totals_row.get(zone_name, 0)
    db_val = db_zone_totals.get(zone_name, 0)
    diff = int(excel_val) - int(db_val) if excel_val and db_val else None
    status = "OK" if excel_val == db_val else f"DIFF: {diff}"
    print(f"  {zone_name}: Excel={excel_val}, DB={db_val} [{status}]")
