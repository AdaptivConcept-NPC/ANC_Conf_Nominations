"""Extract per-ward, per-candidate, per-zone data from BRANCH NOMI."""
import openpyxl
import re
from collections import defaultdict

WARD_PATTERN = re.compile(r"WARD\s*(\d+)", re.IGNORECASE)

wb = openpyxl.load_workbook(r'E:\webdev\ANC_Conf_Nominations\docs\NOM2026 PR and Councillor Nominations.xlsx', data_only=True)
ws = wb['BRANCH NOMI']

# Zone headers in row 3
zone_labels = {}
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=3, column=col_idx).value
    if val and not str(val).strip().lower().startswith('column'):
        zone_labels[col_idx] = str(val).strip()

# Parse: for each zone, for each ward, count candidate occurrences
# Structure: {zone_name: {ward_number: {candidate_name: count}}}
zone_ward_candidates = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

for col_idx, zone_label in zone_labels.items():
    current_ward = None
    for row_idx in range(4, ws.max_row + 1):
        raw_value = str(ws.cell(row=row_idx, column=col_idx).value or "").strip()
        if not raw_value:
            continue

        ward_match = WARD_PATTERN.search(raw_value)
        if ward_match:
            current_ward = int(ward_match.group(1))
            continue

        if current_ward is None:
            continue

        if raw_value.upper().startswith("WARD"):
            continue

        zone_ward_candidates[zone_label][current_ward][raw_value] += 1

# Print structure
for zone in sorted(zone_ward_candidates.keys()):
    print(f"\n=== {zone} ===")
    for ward in sorted(zone_ward_candidates[zone].keys()):
        candidates = zone_ward_candidates[zone][ward]
        total = sum(candidates.values())
        cand_str = ", ".join(f"{n}:{v}" for n, v in sorted(candidates.items()))
        print(f"  Ward {ward} ({total} votes): {cand_str}")
